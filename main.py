import numbers
from os import set_inheritable
import string

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from typing import Tuple, Union
from datetime import date, datetime, datetime

from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File

import pandas as pd 

# For support or further development, contact the author on GitHub or LinkedIn

class hue_proponent:
    cmhc_name: str
    cmhc_id: str
    cmhc_tdsrratioconclusion: Union[float, str]
    cmhc_liquiditycurrentratioconclusion: Union[float, str]
    cmhc_loantovalueratioconclusion: Union[float, str]
    cmhc_dcrratioconclusion: Union[float, str]
    cmhc_overallfinancialcapacityconclusion: Union[float, str]
    cmhc_tierednetworthratioconclusion: Union[float, str]
    cmhc_credithistoryconclusion: Union[float, str]
    
    cmhc_rhccapprovalrequired: bool	
    cmhc_netoperatingincomenoiconso: Union[float, str]	
    cmhc_revenuesconso: Union[float, str]	
    cmhc_outstandingeffectiveexposurerefyear: Union[float, str]		
    cmhc_totaltierednetworthconso: Union[float, str]	
    cmhc_totaltieredliabilitiesconso: Union[float, str]	
    cmhc_adjustedtdsrcashflowconso_base: Union[float, str]
    cmhc_mortgagepayableconso: Union[float, str]	
    cmhc_outstandingeffectiveexposurerefyear_base: Union[float, str]	
    cmhc_netoperatingincomenoiconso_base: Union[float, str]	
    cmhc_adjustedtdsrcashflowconso: Union[float, str]
    cmhc_mortgagepayableconso_base: Union[float, str]	
    mhc_propertyrelatedannualpaymentsconso_base: Union[float, str]	
    cmhc_totaltieredassetsconso: Union[float, str]	
    cmhc_totalannualpaymentsconso_base: Union[float, str]	
    cmhc_incomeproducingpropertiesconso: Union[float, str]	
    cmhc_totalannualpaymentsconso: Union[float, str]	
    cmhc_revenuesconso_base: Union[float, str]	
    cmhc_incomeproducingpropertiesconso_base: Union[float, str]	
    cmhc_propertyrelatedannualpaymentsconso: Union[float, str]	
    cmhc_totaltieredliabilitiesconso_base: Union[float, str]	
    cmhc_totaltierednetworthconso_base: Union[float, str]	
    cmhc_totaltieredassetsconso_base: Union[float, str]	
    cmhc_overallfincreditcapacityconclusion: Union[float, str]	
    cmhc_mitigationactions: Union[float, str]	
    cmhc_tdsry2: Union[float, str]
    cmhc_dcrwithsubs: Union[float, str]	
    cmhc_liquiditycurrentratio: Union[float, str]
    cmhc_referenceyeartext: Union[float, str]
    cmhc_coveragerationwcdrefyear: Union[float, str]	
    cmhc_ltvy1: Union[float, str]	
    cmhc_isdcount: Union[float, str]	
    cmhc_tierednetworthratio: Union[float, str]	
    cmhc_financialreviewsummarycompletiondate: Union[float, str]	
    cmhc_comments: Union[float, str]	
    cmhc_credithistorycomments: Union[float, str]	
    cmhc_currentratioy1: Union[float, str]	
    cmhc_overallrisks: Union[float, str]	
    cmhc_coveragerationwcdy2: Union[float, str]	
    cmhc_tdsrrequirement: Union[float, str]	
    cmhc_tdsrwithsubs: Union[float, str]	
    cmhc_referenceyear: Union[float, str]	
    cmhc_tnwy2: Union[float, str]	
    cmhc_tdsry1: Union[float, str]	
    cmhc_currentratioy2: Union[float, str]	
    cmhc_loantovalueratioy1: Union[float, str]	
    cmhc_loantovalueratioy2: Union[float, str]	
    cmhc_ltvreferenceyear: Union[float, str]	
    cmhc_tdsrreferenceyear: Union[float, str]	
    cmhc_name: Union[float, str]	
    cmhc_loantovalueratio: Union[float, str]	
    cmhc_dcry1: Union[float, str]	
    cmhc_coveragerationwcdy1: Union[float, str]	
    cmhc_dcry2: Union[float, str]	
    cmhc_tnwreferenceyear: Union[float, str]	
    cmhc_ltvwithsubs: Union[float, str]	
    cmhc_dcrreferenceyear: Union[float, str]	
    cmhc_ltvy2: Union[float, str]	
    cmhc_tnwwithsubs: Union[float, str]	
    cmhc_currentratioreferenceyear: Union[float, str]	
    cmhc_bsdcount: Union[float, str]	
    cmhc_tnwy1: Union[float, str]	
    cmhc_hybridtdsr_en: Union[float, str]	
    cmhc_tierednetworthratioconclusion_en: Union[float, str]	
    cmhc_dcrratioconclusion_en: Union[float, str]	
    cmhc_credithistoryconclusion_en: Union[float, str]	
    cmhc_loantovalueratioconclusion_en: Union[float, str]	
    cmhc_liquiditycurrentratioconclusion_en: Union[float, str]	
    cmhc_overallfinancialcapacityconclusion_en: Union[float, str]	
    cmhc_tdsrratioconclusion_en: Union[float, str]

    # data for the Series tab

    caprate: Union[float, str]
    total_units: Union[float, str]
    total_property_value: Union[float, str]
    rental_revenue: Union[float, str]
    residential_rental_revenue: Union[float, str]
    commercial_rental_revenue: Union[float, str]
    adjusted_noi: Union[float, str]
    adjusted_portfolio_debt_servicing_requirements: Union[float, str]
    dcr_data_check: Union[float, str]
    adjusted_total_debt_servicing_requirements: Union[float, str] 
    tdsr_data_check: Union[float, str]
    tdsr_data_check_en: Union[float, str]
    total_assets: Union[float, str]
    total_liabilities: Union[float, str]
    current_ratio_data_check: Union[float, str]
    available_credit_data_check: Union[float, str]
    adjusted_working_capital_surplus: Union[float, str]
    coverage_of_working_capital_data_check: Union[float, str]
    rating_letter: Union[float, str]
    rating_numerical: Union[float, str]
    net_notching_impact_on_scorecard: Union[float, str]
    factor_1_scale: Union[float, str]
    factor_2_business_profile: Union[float, str]
    factor_2_business_profile_operating_environment: Union[float, str] 
    factor_3_liquidity: Union[float, str]
    factor_4_leverage_and_coverage: Union[float, str]
    factor_4_leverage_and_coverage_elt_based: Union[float, str]
    factor_4_leverage_and_coverage_ltv_based: Union[float, str]
    factor_4_leverage_and_coverage_portfolio_dcr_based: Union[float, str]
    factor_4_leverage_and_coverage_tdsr_based: Union[float, str] 
                   
    def __init__(self, row: list) -> None:
        
        self.cmhc_id = row[0]
        self.cmhc_tdsrratioconclusion = row[1]
        self.cmhc_liquiditycurrentratioconclusion = row[2]
        self.cmhc_loantovalueratioconclusion = row[3]
        self.cmhc_dcrratioconclusion = row[4]
        self.cmhc_overallfinancialcapacityconclusion = row[5]
        self.cmhc_tierednetworthratioconclusion = row[6]
        self.cmhc_credithistoryconclusion = row[7]
        
        self.cmhc_rhccapprovalrequiredself = row[8]
        self.cmhc_netoperatingincomenoiconso = row[9]
        self.cmhc_revenuesconso = row[10]	
        self.cmhc_outstandingeffectiveexposurerefyear = row[11]	
        self.cmhc_totaltierednetworthconso = row[12]
        self.cmhc_totaltieredliabilitiesconso = row[13]
        self.cmhc_adjustedtdsrcashflowconso_base = row[14]
        self.cmhc_mortgagepayableconso = row[15]
        self.cmhc_outstandingeffectiveexposurerefyear_base = row[16]
        self.cmhc_netoperatingincomenoiconso_base = row[17]
        self.cmhc_adjustedtdsrcashflowconso = row[18]
        self.cmhc_mortgagepayableconso_base = row[19]
        self.mhc_propertyrelatedannualpaymentsconso_base = row[20]
        self.cmhc_totaltieredassetsconso = row[21]
        self.cmhc_totalannualpaymentsconso_base = row[22]
        self.cmhc_incomeproducingpropertiesconso = row[23]
        self.cmhc_totalannualpaymentsconso = row[24]	
        self.cmhc_revenuesconso_base = row[25]
        self.cmhc_incomeproducingpropertiesconso_base = row[26]
        self.cmhc_propertyrelatedannualpaymentsconso = row[27]
        self.cmhc_totaltieredliabilitiesconso_base = row[28]	
        self.cmhc_totaltierednetworthconso_base = row[29]
        self.cmhc_totaltieredassetsconso_base = row[30]
        self.cmhc_overallfincreditcapacityconclusion = row[31]
        self.cmhc_mitigationactions = row[32]
        self.cmhc_tdsry2 = row[33]
        self.cmhc_dcrwithsubs = row[34]	
        self.cmhc_liquiditycurrentratio = row[35]
        self.cmhc_referenceyeartext = row[36]
        self.cmhc_coveragerationwcdrefyear = row[37]
        self.cmhc_ltvy1 = row[38]	
        self.cmhc_isdcount = row[39]
        self.cmhc_tierednetworthratio = row[40]	
        self.cmhc_financialreviewsummarycompletiondate = row[41]
        self.cmhc_comments = row[42]	
        self.cmhc_credithistorycomments = row[43]
        self.cmhc_currentratioy1 = row[44]
        self.cmhc_overallrisks = row[45]
        self.cmhc_coveragerationwcdy2 = row[46]	
        self.cmhc_tdsrrequirement = row[47]	
        self.cmhc_tdsrwithsubs = row[48]
        self.cmhc_referenceyear = row[49]
        self.cmhc_tnwy2 = row[50]
        self.cmhc_tdsry1 = row[51]
        self.cmhc_currentratioy2 = row[52]
        self.cmhc_loantovalueratioy1 = row[53]
        self.cmhc_loantovalueratioy2 = row[54]
        self.cmhc_ltvreferenceyear = row[55]	
        self.cmhc_tdsrreferenceyear = row[56]	
        self.cmhc_name = row[57]
        self.cmhc_loantovalueratio = row[58]
        self.cmhc_dcry1 = row[59]
        self.cmhc_coveragerationwcdy1 = row[60]	
        self.cmhc_dcry2 = row[61]
        self.cmhc_tnwreferenceyear = row[62]
        self.cmhc_ltvwithsubs = row[63]
        self.cmhc_dcrreferenceyear = row[64]	
        self.cmhc_ltvy2 = row[65]
        self.cmhc_tnwwithsubs = row[66]	
        self.cmhc_currentratioreferenceyear = row[67]
        self.cmhc_bsdcount = row[68]
        self.cmhc_tnwy1 = row[69]
        self.cmhc_hybridtdsr_en = row[70]
        self.cmhc_tierednetworthratioconclusion_en = row[71]	
        self.cmhc_dcrratioconclusion_en = row[72]
        self.cmhc_credithistoryconclusion_en = row[73]
        self.cmhc_loantovalueratioconclusion_en = row[74]
        self.cmhc_liquiditycurrentratioconclusion_en = row[75]
        self.cmhc_overallfinancialcapacityconclusion_en = row[76]	
        self.cmhc_tdsrratioconclusion_en = row[77]

        self.cmhc_name = self.search_for_name(self.cmhc_id, "name_ids.xlsx")
        
    def search_for_name(self, id, filename):
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        for row in sheet.iter_rows():
            if row[0].value == id:
                return row[1].value
        return "Not Found"
        

class proponent_entry(hue_proponent):
    yearvalue: int
    
    def __init__(self, year: int) -> None:
        self.yearvalue = year
        super.__init__()


def initialize_inputs(input_file: str) -> list[hue_proponent]:
    """
    Initialize the inputs from the input file.
    """
    
    book = load_workbook(input_file)
    sheet = book.active
    
    accumulator = []
    
    for row in sheet.iter_rows():
        accumulator.append(hue_proponent([cell.value for cell in row])) 
            
    return accumulator


def initialize_hashmap(input_file_ID_set: str, relational_sheet: str) -> dict[int, str]:
    """
    Initialize the hashmap from the input file.
    
    Preconditions:
    - Input file is an excel spreadsheet with the first column containing the ID
    - Relational sheet is the sheet in the input file that contains the ID and name pairs
    
    
    PAGE LAYOUT SPECIFICATIONS:
    
    FIRST COLUMN: ID
    SECOND COLUMN: NAME
    EVERY ID HAS ASSOCIATED NAME
    """
    
    aw = load_workbook(input_file_ID_set)
    aws = aw.active
    
    bw = load_workbook(relational_sheet)
    bws = bw.active
    
    accdict = {}
    
    # for each ID in the input file, find the name in the relational sheet and
    # append the KVP to the accumulator.
    
    for row in aws.iter_rows():
        rowValues = [cell.value for cell in row]
        accdict[rowValues[0]] = rowValues[1]
    
    for id in input_file_ID_set:
        if id not in accdict:
            accdict[id] = "Not Found"

    return accdict


def return_exact_name_match_ind(input_sheet, target) -> int:
    """
    Iterates through the destination list and returns the index.
    """
        
    for index, sheet in enumerate(input_sheet):
        if sheet.cmhc_name.lower() == target.lower():
            return index
    return -1


def binary_search_match(input_sheet: Workbook, lookfor: hue_proponent) -> int:
    """
    Binary search for the target.
    """
    target = lookfor.cmhc_name
    
    low = 0
    high = len(input_sheet) - 1
    
    while low <= high:
        mid = (low + high) // 2
        if input_sheet['A' + str(mid)].lower() == target.lower():
            # This is the found condition. We want to find the first entry.
            
            if mid <= 2:
                return mid
            else:
                if input_sheet['A' + str(mid - 1)].lower() == target.lower():
                    while input_sheet['A' + str(mid - 1)].lower() == target.lower():
                        mid -= 1
                    return mid
            
            return mid
        elif input_sheet['A' + str(mid)].lower() < target.lower():
            low = mid + 1
        else:
            high = mid - 1
    return -1


def update_series_with_proponent(input_sheet: Workbook, proponent: hue_proponent) -> None:
    '''
    Updates the proponent's entry, taking a workbook and a proponent.
    
    Note: Formatting of Series:
    
    Borrower
    Year
    Last
    LTV	TNW ratio
    TNW without pref shares or prom notes
    DCR	TDSR
    Total TNW
    Total TNW without pref shares or prom notes
    Total NOI
    Cap Rate
    Total Number of Properties
    Total Property Value
    Rental Revenue
    Residential Rental Revenue (From IS)
    Commercial Rental Revenue (From IS)
    Adjusted NOI
    Adjusted Portfolio Debt Servicing Requirements
    DCR data Check
    Adjusted Total Debt Servicing Requirements
    Cashflow for debt servicing (TDSR Cashflow)
    TDSR Data Check	Current Assets
    Total Assets (from BS)
    Current Liabilities	Total Liabilities (from BS)
    Current Ratio Data Check
    Available Credit (Unused Credit Facility)
    Adjusted working capital surplus
    Coverage of Working Capital	Rating (Letter)
    Rating (Numerical)
    Net Notching Impact on Scorecard
    Factor 1 : Scale (Total tiered asset)
    Factor 2 : Business Profile (Market positioning & Asset quality)
    Factor 2 : Business Profile (Operating environment)
    Factor 3 : Liquidity (Liquidity & Access to capital)
    Factor 4 : Leverage & Coverage (Overall leverage)
    Factor 4 : Leverage & Coverage (elt b)		
    Factor 4 : Leverage & Coverage (LTV)		
    Factor 4 : Leverage & Coverage (Portfolio DCR)		
    Factor 4 : Leverage & Coverage (TDSR)					
    # of Notching		
    # From	
    # To	
    # Score after notching	
    # Check																					

    '''
    # open worksheet
    sheet = input_sheet.active
    
    #navigate to correct row (locate index)
    #modify all cells in the numbered entry column by one
    
    index = binary_search_match(input_sheet, proponent)
    
    cell_at_index = sheet['C' + str(index)]
    
    q = int(index)
    c = 1
    
    while sheet['A' + str(q)].value == proponent.cmhc_name:
        sheet['C' + str(q)] = sheet['C' + str(q)].value + 1
        sheet['D' + str(q)] = (str(sheet['D' + str(q)].value))[:-1] + (int(sheet['D' + str(q)].value[-1]) + 1)
        
        q += 1
    
    #then insert the modified row into the sheet
    # also need to update year count
    
    modifiedrow = [proponent.cmhc_name, proponent.cmhc_referenceyear, 1, proponent.cmhc_name + str(1), 
                   "Placeholder for formula", proponent.cmhc_ltvy1, proponent.cmhc_tnwy1, proponent.cmhc_tnwwithsubs, 
                   proponent.cmhc_dcrratioconclusion, proponent.cmhc_tdsrratioconclusion, proponent.cmhc_totaltierednetworthconso, 
                   proponent.cmhc_totaltierednetworthconso_base, proponent.cmhc_netoperatingincomenoiconso, proponent.caprate, 
                   proponent.total_units, proponent.total_property_value, proponent.rental_revenue, proponent.residential_rental_revenue, 
                   proponent.commercial_rental_revenue, proponent.adjusted_noi, proponent.adjusted_portfolio_debt_servicing_requirements, 
                   proponent.dcr_data_check, proponent.adjusted_total_debt_servicing_requirements, proponent.tdsr, proponent.tdsr_data_check, 
                   proponent.total_assets, proponent.total_liabilities, proponent.current_ratio_data_check, proponent.available_credit, proponent.adjusted_working_capital_surplus, 
                   proponent.coverage_of_working_capital, proponent.rating_letter, proponent.rating_numerical, proponent.net_notching_impact_on_scorecard, 
                   proponent.factor_1_scale, proponent.factor_2_business_profile, proponent.factor_2_business_profile_operating_environment, proponent.factor_3_liquidity, 
                   proponent.factor_4_leverage_and_coverage, proponent.factor_4_leverage_and_coverage_elt_b, proponent.factor_4_leverage_and_coverage_ltv, 
                   proponent.factor_4_leverage_and_coverage_portfolio_dcr, proponent.factor_4_leverage_and_coverage_tdsr, proponent.number_of_notching, 
                   proponent.from_, proponent.to, proponent.score_after_notching, proponent.check]
    
    # TODO: Add the new row to the sheet
    
    sheet.insert_rows(index, 1)
    
    # insert modified row into the row at index
    
    for i in range(len(modifiedrow)):
        sheet.cell(row=index, column=i+1, value=modifiedrow[i])
    
    return None
    
    
    

def initialize_hue_proponents(input_file: str) -> list[hue_proponent]:
    """
    Initialize the inputs from the input file.
    """
    
    book = load_workbook(input_file)
    sheet = book.active
    
    accumulator = []
    
    for row in sheet.iter_rows():
        accumulator.append(hue_proponent([cell.value for cell in row])) 
            
    return accumulator

def main(dataset: str, existing_db: str, output_file: str) -> None:
    '''This function contains most of the functionality. It loads input and database files into memory, 
        then after editing saves them into the designated output location. 
    '''
    # Load input files
    wb_input = openpyxl.load_workbook(dataset, data_only=True)
    wb_db = openpyxl.load_workbook(existing_db, data_only=True)
    wb_output_file = openpyxl.load_workbook(output_file, data_only=True)
    
    # Load worksheets
    input_sheet = wb_input.active
    database_sheet = wb_db['Series']
    output_sheet = wb_output_file.active
    
    # Initialize the proponents
    propList = initialize_hue_proponents(dataset)
    
    # ID dictionary
    id_dict = initialize_hashmap(input_sheet)
    
    for proponent in propList:
        print(proponent.proponent_name)
        print(proponent.exposure_time)
        print(proponent.last_exposure)
        print(proponent.last_financial_review_date)
        print(proponent.fiscal_year_end)
        print(proponent.active_FCR_expiry)
        print(proponent.status)
        print(proponent.notes)
        print("\n")
        
        index = binary_search_match(database_sheet, proponent.proponent_name)
    # Update the database with each value in the input file
    

    for row in input_sheet.iter_rows():
        # get the name value from the first cell
        prop_name = row[0].value
        
        if prop_name is not None:

            # turn the row into a proponent object
            
            temp_proponent = proponent([cell.value for cell in row])
            
            # search the database sheet for the row with the same name
            
            index = binary_search_match(database_sheet, prop_name)
            
            # Insert the newest entry into the database, but ONLY if the date
            # of the entry is newer than the existing entry.
            # update those into the db
            
            if temp_proponent.cmhc_referenceyear > row['B' + str(index)].value:
                database_sheet.insert_rows(index, 1)
            
            update_series_with_proponent(database_sheet, temp_proponent)
            # insert the updated entry above the index
            
    
    # Save output files
    wb_output_file.save(output_file)
    wb_db.save(existing_db)
    return None
    