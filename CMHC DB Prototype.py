"""
    Purpose:
    - This is a prototype for the CMHC database made by Kevin Chen
    - The goal is to effectively retrieve information from several external sources
    and then to integrate them into a single database presented in CSV format
    - The resulting data should be able to be easily imported and edited manually
    and the application will not overwrite the data manually added by the user
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from typing import Tuple, Union
from datetime import date, datetime, datetime

workbook = load_workbook(filename='CMHC Database.xlsx')

sheet = workbook.active

class borrower:
    # The second string in each tuple is the cell's color
    # red = "FF0000"
    # green = "00FF00"
    # yellow = "FFFF00"
    
    """
    
    Data tab
    
    var_FS_High: Tuple(str, str)
    var_FS_Low: Tuple(str, str)
    
    var_Last_FR_Date: datetime.date
    var_Active_Fr_Expiry: datetime.date
    var_Fiscal_Year_End: datetime.date
    
    var_LTV: float
    var_TNW_ratio: float
    var_TNW_nopref_ratio: float
    var_DCR: float
    var_TDSR: float
    var_TNW_total: int
    var_TNW_nopref_total: int
    var_NOI_total: int
    var_cap_rate: float
    var_total_number_properties: int
    var_total_property_value: int
    
    var_rental_revenue: int
    var_residental_rental_revenue: int
    var_commercial_rental_revenue: Union(int, str)
    var_adj_NOI: int
    var_adj_portfolio_DSR: int
    
    var_DCR: float # Debt-to-Capital Ratio is NOI / DS
    # The DCR data check is 1 - var_DCR
    ...
    """
    rating: tuple() # The borrower's rating. 
    proponent_name: str # The borrower's proponent name
    
    exposure_time: datetime.date
    last_exposure: int # The borrower's exposure at the exposure_time
    
    last_financial_review_date: datetime.date
    fiscal_year_end: datetime.date
    
    active_FCR_expiry: datetime.date
    
    status: str 
    # The borrower's status, a short custom blurb. Should be empty in the 
    # machine-generated spreadsheet
    
    notes: str
    # Notes on the borrower. Should not be changed by this script.
    
    def __init__(self, rating: str, proponent_name: str, exposure_time: str, last_exposure: int, last_financial_review_date: str, fiscal_year_end: str, active_FCR_expiry: str, status: str, notes: str):
        color_map = {"yellow": "FFFF00", "green": "00FF00", "red": "FF0000"}
        self.rating = (rating, color_map[rating.lower()])
        
        self.proponent_name = int(proponent_name)
        self.exposure_time = self.convert_datetime(exposure_time)
        self.last_exposure = last_exposure
        
        self.last_financial_review_date = self.convert_datetime(last_financial_review_date)
        self.fiscal_year_end = self.convert_datetime(fiscal_year_end)
        self.active_FCR_expiry = self.convert_datetime(active_FCR_expiry)
        
        self.status = status
        self.notes = notes
        
        
    def convert_datetime(self, datetime_str: str) -> datetime.date:
        """
        Converts a datetime string to a datetime.date object
        
        Takes in a string of the form "MM/DD/YYYY"
        """
        return_date = datetime.strptime(datetime_str, "%m/%d/%Y")
        return return_date
    
    def render_row(self, sheet, pos_x_char: int, pos_y: int) -> None:
        """
        Adds the borrower's data to the spreadsheet in the specific row and column
        """
        sheet[pos_x_char + str(pos_y)] = self.rating[0]
        sheet[pos_x_char + str(pos_y)].fill = PatternFill(fgColor=self.rating[1], fill_type="solid")
        
        sheet[chr(ord(pos_x_char) + 1) + str(pos_y)] = self.proponent_name
        sheet[chr(ord(pos_x_char) + 2) + str(pos_y)] = self.last_exposure
        sheet[chr(ord(pos_x_char) + 3) + str(pos_y)] = self.last_financial_review_date.strftime("%m/%d/%Y")
        sheet[chr(ord(pos_x_char) + 4) + str(pos_y)] = self.fiscal_year_end.strftime("%m/%d/%Y")
        sheet[chr(ord(pos_x_char) + 5) + str(pos_y)] = self.active_FCR_expiry.strftime("%m/%d/%Y")
        sheet[chr(ord(pos_x_char) + 6) + str(pos_y)] = self.status
        
        
        
        
        
class spreadsheet_base:
    borrower_list: list[borrower]
    
    def __init__(self, borrowers: list[borrower]):
        self.borrower_list = borrowers

    def sort_by_name(self):
        self.borrower_list.sort(key=lambda x: x.proponent_name)
        
    def render_spreadsheet(self):
        for i in range(2, len(self.borrower_list) + 2):
            self.borrower_list[i - 2].render_row(sheet, "A", i)
            
    def refresh_status_and_notes(self):
        for i in range(2, len(self.borrower_list) + 2):
            self.borrower_list[i - 2].status = sheet["G" + str(i)].value
            self.borrower_list[i - 2].notes = sheet["H" + str(i)].value
            