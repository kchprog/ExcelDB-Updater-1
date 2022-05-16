"""
    Purpose:
    - This is a prototype for the CMHC database made by Kevin Chen
    - The goal is to effectively retrieve information from several external sources
    and then to integrate them into a single database presented in CSV format
    - The resulting data should be able to be easily imported and edited manually
    and the application will not overwrite the data manually added by the user
"""

from email import contentmanager
from functools import total_ordering
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from typing import Tuple, Union
from datetime import date, datetime, datetime

from office365.runtime.auth.authentication_context import AuthenticationContext
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File

import os

'''
Approach: Use the Excel API to directly access and manipulate the data on the online spreadsheet.
'''

def main():
    url = 'https://cmhcschl-my.sharepoint.com/personal/kchen_cmhc-schl_gc_ca'
    relative_url = '/personal/kchen_cmhc-schl_gc_ca/Documents/Database%20Revamp%20Project%20-%20API%20Test%20Storage.xlsx'

    username = 'USERNAME'
    password = 'PASSWORD'


    ctx_auth = AuthenticationContext(url)
    if ctx_auth.acquire_token_for_user(username, password):
        ctx = ClientContext(url, ctx_auth)
        web = ctx.web
        ctx.load(web)
        ctx.execute_query()
        print("Authentication successful")
    else:
        print("Authentication failed")
    # Test: Download file, open, edit, and reupload
    # Download file

    filename = 'Database Revamp Project - API Test Storage.xlsx'
    with open(filename, 'wb') as output_file:
        response = File.open_binary(ctx, relative_url)
        output_file.write(response.content)
        
    # Open file
    wb = load_workbook(filename)
    print(wb.get_named_ranges())
    
    ws = wb.active
    
    print(ws['A1'].value)
    
    ws['B1'] = 'Hello World'
    
    import datetime
    
    ws['B2'] = datetime.datetime.now()
    
    print(ws['B1'].value)
    print(ws['B2'].value)
    
    
    wb.save(filename)
    
    # attempt to upload file back to the sharepoint
    
    with open(filename, 'rb') as input_file:
        file_content = input_file.read()
    
    target_folder = web.get_folder_by_server_relative_url("/personal/kchen_cmhc-schl_gc_ca/Documents/")
    name = os.path.basename(filename)
    target_file = target_folder.upload_file(name, file_content).execute_query()
    
    print("File has been uploaded to url: {0}".format(target_file.serverRelativeUrl))
    
if __name__ == '__main__':
    main()
